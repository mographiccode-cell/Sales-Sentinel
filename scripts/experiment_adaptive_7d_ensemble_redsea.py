from __future__ import annotations

import json
import math
import os
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from app.services.adaptive_forecasting_engine import _CANDIDATES, _select, _simulate, forecast

REPORT_DIR = Path("reports/experiment_adaptive_7d_ensemble_redsea")


def _as_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Unsupported date value: {value!r}")


def load_daily_sales(path: Path) -> tuple[list[date], list[float]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(v).strip() if v is not None else "" for v in next(rows)]
    index = {name: i for i, name in enumerate(header)}
    date_idx = index["TRX DATE"]
    net_idx = index["Net Amount"]
    daily: dict[date, float] = {}
    for row in rows:
        if not row or row[date_idx] in (None, ""):
            continue
        day = _as_date(row[date_idx])
        try:
            value = float(row[net_idx] or 0.0)
        except (TypeError, ValueError):
            continue
        daily[day] = daily.get(day, 0.0) + value
    wb.close()
    start, end = min(daily), max(daily)
    dates, values = [], []
    cursor = start
    while cursor <= end:
        dates.append(cursor)
        values.append(max(0.0, daily.get(cursor, 0.0)))
        cursor += timedelta(days=1)
    return dates, values


def _weighted(paths: dict[str, list[float]], errors: dict[str, float], *, top_k: int | None, power: float, floor: float) -> list[float]:
    eligible = [(name, error) for name, error in errors.items() if math.isfinite(error)]
    eligible.sort(key=lambda item: (item[1], item[0]))
    if top_k is not None:
        eligible = eligible[:top_k]
    raw = {name: 1.0 / ((error + floor) ** power) for name, error in eligible}
    total = sum(raw.values()) or 1.0
    weights = {name: value / total for name, value in raw.items()}
    return [sum(weights[name] * paths[name][i] for name in weights) for i in range(7)]


def _median_path(paths: dict[str, list[float]], errors: dict[str, float], top_k: int) -> list[float]:
    names = [name for name, error in sorted(errors.items(), key=lambda item: (item[1], item[0])) if math.isfinite(error)][:top_k]
    out = []
    for i in range(7):
        vals = sorted(paths[name][i] for name in names)
        mid = len(vals) // 2
        out.append(vals[mid] if len(vals) % 2 else (vals[mid - 1] + vals[mid]) / 2.0)
    return out


def _trimmed_mean(paths: dict[str, list[float]], errors: dict[str, float]) -> list[float]:
    names = [name for name, error in errors.items() if math.isfinite(error)]
    out = []
    for i in range(7):
        vals = sorted(paths[name][i] for name in names)
        trimmed = vals[1:-1] if len(vals) > 4 else vals
        out.append(sum(trimmed) / len(trimmed))
    return out


def evaluate(dates: list[date], values: list[float]) -> dict:
    totals_actual: list[float] = []
    totals_by_strategy: dict[str, list[float]] = {}
    folds: list[dict] = []

    for origin in range(56, len(values) - 7 + 1, 7):
        train = values[:origin]
        actual = values[origin: origin + 7]
        actual_total = sum(actual)
        base_generated = forecast(train, dates[origin - 1], 7)
        base_path = [float(row["predicted"]) for row in base_generated]

        _selected, _selected_metrics, diagnostics = _select(train, 7)
        paths = {name: _simulate(train, name, 7) for name in _CANDIDATES}
        errors = {
            name: (float(result["total_wape"]) if math.isfinite(float(result["total_wape"])) else float(result["wape"]))
            for name, result in diagnostics.items()
        }

        strategies = {
            "v3_winner": base_path,
            "inverse_all_p1": _weighted(paths, errors, top_k=None, power=1.0, floor=0.05),
            "inverse_all_p2": _weighted(paths, errors, top_k=None, power=2.0, floor=0.05),
            "inverse_top3_p1": _weighted(paths, errors, top_k=3, power=1.0, floor=0.05),
            "inverse_top3_p2": _weighted(paths, errors, top_k=3, power=2.0, floor=0.05),
            "inverse_top5_p1": _weighted(paths, errors, top_k=5, power=1.0, floor=0.05),
            "median_top3": _median_path(paths, errors, 3),
            "median_top5": _median_path(paths, errors, 5),
            "trimmed_mean_all": _trimmed_mean(paths, errors),
        }
        totals_actual.append(actual_total)
        fold = {"origin": dates[origin].isoformat(), "actual_total": actual_total, "v3_model": base_generated[0]["model_name"]}
        for name, path in strategies.items():
            pred_total = sum(path)
            totals_by_strategy.setdefault(name, []).append(pred_total)
            fold[name] = {
                "predicted_total": pred_total,
                "abs_error_pct": abs(actual_total - pred_total) / max(abs(actual_total), 1e-9) * 100.0,
            }
        folds.append(fold)

    denom = sum(abs(v) for v in totals_actual)
    summary = {}
    for name, predictions in totals_by_strategy.items():
        wape = sum(abs(a - p) for a, p in zip(totals_actual, predictions)) / max(denom, 1e-9)
        summary[name] = {"horizon_total_wape": wape, "quality_proxy_pct": max(0.0, (1.0 - wape) * 100.0)}
    ranked = sorted(summary.items(), key=lambda item: item[1]["horizon_total_wape"])
    return {"fold_count": len(folds), "ranked": ranked, "summary": summary, "folds": folds}


def main() -> None:
    source = Path(os.environ.get("REDSEA_XLSX", "/tmp/RedSea_Data_Cleaned.xlsx"))
    dates, values = load_daily_sales(source)
    result = evaluate(dates, values)
    payload = {
        "status": "POST_OPEN_DEVELOPMENT_EXPERIMENT_NOT_BLIND_VALIDATION",
        "purpose": "Leakage-safe nested weighting experiment for the 7-day horizon; outer Redsea folds are never used to compute a fold's ensemble weights.",
        "calendar_days": len(values),
        "date_start": dates[0].isoformat(),
        "date_end": dates[-1].isoformat(),
        **result,
    }
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    (REPORT_DIR / "metrics.json").write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps({"status": payload["status"], "fold_count": result["fold_count"], "ranked": result["ranked"]}, indent=2))


if __name__ == "__main__":
    main()
