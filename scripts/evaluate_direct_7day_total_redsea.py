from __future__ import annotations

import json
import os
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

SOURCE_SHA256 = "dd994d25042e54119babe058820b02e00b3443fe18150b9213f9c6929466a645"
REPORT_DIR = Path("reports/external_redsea_direct_7day_total")
CANDIDATES = (
    "last_7d_total",
    "mean_2w_total",
    "mean_4w_total",
    "mean_8w_total",
    "median_4w_total",
    "median_8w_total",
    "trimmed_mean_8w_total",
    "ewma_8w_total",
    "blend_mean_median_8w",
    "blend_mean8_recent",
    "weighted_4w_total",
    "damped_weekly_trend",
)


def _as_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Unsupported date: {value!r}")


def load_daily(path: Path) -> tuple[list[date], list[float], int]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(v).strip() if v is not None else "" for v in next(rows)]
    idx = {name: i for i, name in enumerate(header)}
    di, ni = idx.get("TRX DATE"), idx.get("Net Amount")
    if di is None or ni is None:
        raise RuntimeError("Required Redsea columns are missing")
    daily: dict[date, float] = {}
    source_rows = 0
    for row in rows:
        if not row or row[di] in (None, ""):
            continue
        source_rows += 1
        day = _as_date(row[di])
        try:
            net = float(row[ni] or 0.0)
        except (TypeError, ValueError):
            continue
        daily[day] = daily.get(day, 0.0) + net
    wb.close()
    start, end = min(daily), max(daily)
    dates, values = [], []
    cur = start
    while cur <= end:
        dates.append(cur)
        values.append(max(0.0, daily.get(cur, 0.0)))
        cur += timedelta(days=1)
    return dates, values, source_rows


def _median(values: list[float]) -> float:
    ordered = sorted(values)
    if not ordered:
        return 0.0
    m = len(ordered) // 2
    return ordered[m] if len(ordered) % 2 else (ordered[m - 1] + ordered[m]) / 2.0


def _mean(values: list[float]) -> float:
    return sum(values) / len(values) if values else 0.0


def _weekly_totals(history: list[float], weeks: int = 8) -> list[float]:
    totals = []
    for w in range(weeks):
        end = len(history) - 7 * w
        start = end - 7
        if start < 0:
            break
        totals.append(sum(history[start:end]))
    return totals


def predict_total(history: list[float], model: str) -> float:
    weeks = _weekly_totals(history, 8)
    if not weeks:
        return 0.0
    if model == "last_7d_total":
        value = weeks[0]
    elif model == "mean_2w_total":
        value = _mean(weeks[:2])
    elif model == "mean_4w_total":
        value = _mean(weeks[:4])
    elif model == "mean_8w_total":
        value = _mean(weeks[:8])
    elif model == "median_4w_total":
        value = _median(weeks[:4])
    elif model == "median_8w_total":
        value = _median(weeks[:8])
    elif model == "trimmed_mean_8w_total":
        subset = sorted(weeks[:8])
        trimmed = subset[1:-1] if len(subset) >= 5 else subset
        value = _mean(trimmed)
    elif model == "ewma_8w_total":
        subset = weeks[:8]
        weights = [0.34, 0.23, 0.16, 0.11, 0.07, 0.05, 0.025, 0.015][:len(subset)]
        value = sum(v * w for v, w in zip(subset, weights)) / sum(weights)
    elif model == "blend_mean_median_8w":
        subset = weeks[:8]
        value = 0.55 * _mean(subset) + 0.45 * _median(subset)
    elif model == "blend_mean8_recent":
        subset = weeks[:8]
        value = 0.70 * _mean(subset) + 0.30 * weeks[0]
    elif model == "weighted_4w_total":
        subset = weeks[:4]
        weights = [0.40, 0.30, 0.20, 0.10][:len(subset)]
        value = sum(v * w for v, w in zip(subset, weights)) / sum(weights)
    elif model == "damped_weekly_trend":
        if len(weeks) < 2:
            value = weeks[0]
        else:
            raw = weeks[0] + 0.45 * (weeks[0] - weeks[1])
            value = min(weeks[0] * 1.30, max(weeks[0] * 0.70, raw))
    else:
        raise ValueError(model)
    return max(0.0, float(value))


def _local_score(history: list[float], model: str) -> tuple[float, int]:
    latest = len(history) - 7
    if latest < 28:
        return float("inf"), 0
    start = max(28, latest - 56)
    actuals, preds = [], []
    for origin in range(start, latest + 1, 7):
        train = history[:origin]
        actuals.append(sum(history[origin:origin + 7]))
        preds.append(predict_total(train, model))
    error = sum(abs(a - p) for a, p in zip(actuals, preds))
    denom = sum(abs(a) for a in actuals)
    return error / max(denom, 1e-9), len(actuals)


def select(history: list[float]) -> tuple[str, dict[str, float], int]:
    scores: dict[str, float] = {}
    folds = 0
    for model in CANDIDATES:
        score, n = _local_score(history, model)
        scores[model] = score
        folds = max(folds, n)
    selected = min(CANDIDATES, key=lambda m: (scores[m], m))
    return selected, scores, folds


def weekday_shares(history: list[float]) -> list[float]:
    levels = []
    for future_pos in range(7):
        samples = []
        first_lag = 7 - future_pos
        for lag in range(first_lag, min(len(history), first_lag + 56) + 1, 7):
            samples.append(history[-lag])
        levels.append(_median(samples) if samples else 1.0)
    total = sum(levels)
    return [v / total for v in levels] if total > 0 else [1.0 / 7.0] * 7


def evaluate(dates: list[date], values: list[float]) -> dict:
    actuals, preds = [], []
    winners: Counter[str] = Counter()
    folds = []
    for origin in range(56, len(values) - 7 + 1, 7):
        history = values[:origin]
        actual_total = sum(values[origin:origin + 7])
        winner, scores, local_folds = select(history)
        predicted_total = predict_total(history, winner)
        shares = weekday_shares(history)
        daily_path = [predicted_total * share for share in shares]
        actuals.append(actual_total)
        preds.append(predicted_total)
        winners[winner] += 1
        folds.append({
            "origin": dates[origin].isoformat(),
            "winner": winner,
            "local_backtest_folds": local_folds,
            "candidate_total_wape": scores,
            "actual_total": actual_total,
            "predicted_total": predicted_total,
            "absolute_error_pct": abs(actual_total - predicted_total) / max(abs(actual_total), 1e-9) * 100.0,
            "daily_path": daily_path,
        })
    error = sum(abs(a - p) for a, p in zip(actuals, preds))
    denom = sum(abs(a) for a in actuals)
    wape = error / max(denom, 1e-9)
    return {
        "fold_count": len(folds),
        "total_wape": wape,
        "quality_proxy_1_minus_wape": max(0.0, 1.0 - wape),
        "current_v3_total_wape_reference": 0.3565,
        "improvement_vs_v3_pct_points": (0.3565 - wape) * 100.0,
        "winner_counts": dict(winners),
        "folds": folds,
    }


def main() -> None:
    source = Path(os.environ.get("REDSEA_XLSX", "/tmp/RedSea_Data_Cleaned.xlsx"))
    dates, values, source_rows = load_daily(source)
    result = evaluate(dates, values)
    payload = {
        "experiment": "Direct 7-day total forecast with robust 8-week candidate pool and weekday-share decomposition",
        "scientific_status": "POST_OPEN_EXTERNAL_EXPERIMENT_NOT_FRESH_BLIND",
        "source_sha256": SOURCE_SHA256,
        "source_rows": source_rows,
        "calendar_days": len(values),
        "date_start": dates[0].isoformat(),
        "date_end": dates[-1].isoformat(),
        "candidate_models": list(CANDIDATES),
        "result": result,
        "acceptance_rule": "Do not replace V3 unless total WAPE is lower than V3 and full runtime tests pass.",
    }
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    (REPORT_DIR / "metrics.json").write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    summary = f"""# Direct 7-day total forecast — Redsea experiment

- Source rows: **{source_rows}**
- Calendar days: **{len(values)}** ({dates[0]} to {dates[-1]})
- Candidate models: **{len(CANDIDATES)}**
- Folds: **{result['fold_count']}**
- Direct-total WAPE: **{result['total_wape'] * 100:.2f}%**
- Direct-total quality proxy (1-WAPE): **{result['quality_proxy_1_minus_wape'] * 100:.2f}%**
- Current V3 7-day total WAPE reference: **35.65%**
- Improvement vs V3: **{result['improvement_vs_v3_pct_points']:.2f} percentage points**
- Winner counts: `{json.dumps(result['winner_counts'], ensure_ascii=False)}`

## Decision rule

This remains a post-open Redsea experiment. It may replace the V3 seven-day point-forecast path only if it improves total WAPE and then passes the full application CI. It is not fresh blind validation.
"""
    (REPORT_DIR / "summary.md").write_text(summary, encoding="utf-8")
    print(summary)


if __name__ == "__main__":
    main()
