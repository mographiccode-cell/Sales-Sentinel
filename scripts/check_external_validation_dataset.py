from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from openpyxl import load_workbook

from app.services.external_validation_gate import assess_external_merchant_rows


def _read_csv(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        return list(csv.DictReader(stream))


def _read_xlsx(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value).strip() if value is not None else "" for value in next(iterator)]
    except StopIteration:
        workbook.close()
        return []
    rows = [dict(zip(headers, values)) for values in iterator]
    workbook.close()
    return rows


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Check whether a fresh merchant dataset is suitable for blind Sales Sentinel validation before any model tuning."
    )
    parser.add_argument("path", type=Path)
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()

    suffix = args.path.suffix.lower()
    if suffix == ".csv":
        rows = _read_csv(args.path)
    elif suffix == ".xlsx":
        rows = _read_xlsx(args.path)
    else:
        raise SystemExit("Only CSV and XLSX are supported")

    result = assess_external_merchant_rows(rows).to_dict()
    payload = json.dumps(result, indent=2, ensure_ascii=False)
    print(payload)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(payload + "\n", encoding="utf-8")
    return 0 if result["eligible"] else 2


if __name__ == "__main__":
    raise SystemExit(main())
