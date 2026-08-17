from __future__ import annotations

import json
import math
import os
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from app.services.adaptive_forecasting_engine import forecast

REPORT_DIR = Path("reports/experiment_adaptive_7d_bias_redsea")


def _as_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Unsupported date value: {value!r}")


def load_daily_sales(path: Path) -> tuple[list[date], list[float]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(v).strip() if v is not None else "" for v in next(rows)]
    index = {name: i for i, name in enumerate(header)}
    date_idx = index["TRX DATE"]
    net_idx = index["Net Amount"]
    daily: dict[date, float] = {}
    for row in rows:
        if not row or row[date_idx] in (None, ""):
            continue
        day = _as_date(row[date_idx])
        try:
            value = float(row[net_idx] or 0.0)
        except (TypeError, ValueError):
            continue
        daily[day] = daily.get(day, 0.0) + value
    wb.close()
    start, end = min(daily), max(daily)
    dates, values = [], []
    cursor = start
    while cursor <= end:
        dates.append(cursor)
        values.append(max(0.0, daily.get(cursor, 0.0)))
        cursor += timedelta(days=1)
    return dates, values


def _v3_path(history: list[float], last_date: date) -> list[float]:
    return [float(row["predicted"]) for row in forecast(history, last_date, 7)]


def _prior_ratios(values: list[float], dates: list[date], origin: int, *, max_folds: int = 8) -> list[float]:
    # Construct only completed weekly forecasts strictly before the outer origin.
    inner_origins = list(range(56, origin - 7 + 1, 7))
    inner_origins = inner_origins[-max_folds:]
    ratios: list[float] = []
    for inner in inner_origins:
        history = values[:inner]
        predicted_total = sum(_v3_path(history, dates[inner - 1]))
        actual_total = sum(values[inner:inner + 7])
        if predicted_total > 1e-9 and math.isfinite(predicted_total) and math.isfinite(actual_total):
            ratios.append(actual_total / predicted_total)
    return ratios


def _median(values: list[float]) -> float:
    ordered = sorted(values)
    if not ordered:
        return 1.0
    mid = len(ordered) // 2
    return ordered[mid] if len(ordered) % 2 else (ordered[mid - 1] + ordered[mid]) / 2.0


def _mean(values: list[float]) -> float:
    return sum(values) / len(values) if values else 1.0


def _trimmed(values: list[float]) -> float:
    ordered = sorted(values)
    if len(ordered) >= 5:
        ordered = ordered[1:-1]
    return _mean(ordered)


def _ewma(values: list[float], alpha: float) -> float:
    if not values:
        return 1.0
    level = values[0]
    for value in values[1:]:
        level = alpha * value + (1.0 - alpha) * level
    return level


def _clip(value: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, value))


def evaluate(dates: list[date], values: list[float]) -> dict:
    actual_totals: list[float] = []
    predicted: dict[str, list[float]] = {}
    folds: list[dict] = []
    for origin in range(56, len(values) - 7 + 1, 7):
        history = values[:origin]
        base_path = _v3_path(history, dates[origin - 1])
        base_total = sum(base_path)
        actual_total = sum(values[origin:origin + 7])
        ratios8 = _prior_ratios(values, dates, origin, max_folds=8)
        ratios4 = ratios8[-4:]
        ratios6 = ratios8[-6:]

        factors = {
            "v3_uncorrected": 1.0,
            "median4_clip_080_120": _clip(_median(ratios4), 0.80, 1.20),
            "median6_clip_080_120": _clip(_median(ratios6), 0.80, 1.20),
            "median8_clip_080_120": _clip(_median(ratios8), 0.80, 1.20),
            "mean4_clip_080_120": _clip(_mean(ratios4), 0.80, 1.20),
            "trimmed8_clip_080_120": _clip(_trimmed(ratios8), 0.80, 1.20),
            "ewma8_a035_clip_080_120": _clip(_ewma(ratios8, 0.35), 0.80, 1.20),
            "ewma8_a050_clip_080_120": _clip(_ewma(ratios8, 0.50), 0.80, 1.20),
            "median8_clip_075_125": _clip(_median(ratios8), 0.75, 1.25),
            "ewma8_a035_clip_075_125": _clip(_ewma(ratios8, 0.35), 0.75, 1.25),
        }
        actual_totals.append(actual_total)
        fold = {
            "origin": dates[origin].isoformat(),
            "actual_total": actual_total,
            "base_total": base_total,
            "prior_ratio_count": len(ratios8),
            "prior_ratios": ratios8,
        }
        for name, factor in factors.items():
            pred = base_total * factor
            predicted.setdefault(name, []).append(pred)
            fold[name] = {
                "factor": factor,
                "predicted_total": pred,
                "abs_error_pct": abs(actual_total - pred) / max(abs(actual_total), 1e-9) * 100.0,
            }
        folds.append(fold)

    denom = sum(abs(v) for v in actual_totals)
    summary = {}
    for name, preds in predicted.items():
        wape = sum(abs(a - p) for a, p in zip(actual_totals, preds)) / max(denom, 1e-9)
        summary[name] = {
            "horizon_total_wape": wape,
            "quality_proxy_pct": max(0.0, (1.0 - wape) * 100.0),
        }
    ranked = sorted(summary.items(), key=lambda item: item[1]["horizon_total_wape"])
    return {"fold_count": len(folds), "ranked": ranked, "summary": summary, "folds": folds}


def main() -> None:
    source = Path(os.environ.get("REDSEA_XLSX", "/tmp/RedSea_Data_Cleaned.xlsx"))
    dates, values = load_daily_sales(source)
    result = evaluate(dates, values)
    payload = {
        "status": "POST_OPEN_DEVELOPMENT_EXPERIMENT_NOT_BLIND_VALIDATION",
        "purpose": "Leakage-safe weekly level-bias correction using only completed prior V3 weekly forecast errors.",
        "calendar_days": len(values),
        "date_start": dates[0].isoformat(),
        "date_end": dates[-1].isoformat(),
        **result,
    }
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    (REPORT_DIR / "metrics.json").write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps({"status": payload["status"], "fold_count": result["fold_count"], "ranked": result["ranked"]}, indent=2))


if __name__ == "__main__":
    main()
