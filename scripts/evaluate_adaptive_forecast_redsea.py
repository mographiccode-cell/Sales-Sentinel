from __future__ import annotations

import json
import math
import os
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from app.services.adaptive_forecasting_engine import MODEL_VERSION, forecast

SOURCE_SHA256 = "dd994d25042e54119babe058820b02e00b3443fe18150b9213f9c6929466a645"
REPORT_DIR = Path("reports/external_redsea_adaptive_forecast")


def _as_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Unsupported date value: {value!r}")


def load_daily_sales(path: Path) -> tuple[list[date], list[float], int]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(v).strip() if v is not None else "" for v in next(rows)]
    index = {name: i for i, name in enumerate(header)}
    date_idx = index.get("TRX DATE")
    net_idx = index.get("Net Amount")
    if date_idx is None or net_idx is None:
        raise RuntimeError(f"Required Redsea columns missing. Found: {header}")

    daily: dict[date, float] = {}
    source_rows = 0
    for row in rows:
        if not row or row[date_idx] in (None, ""):
            continue
        source_rows += 1
        day = _as_date(row[date_idx])
        try:
            value = float(row[net_idx] or 0.0)
        except (TypeError, ValueError):
            continue
        daily[day] = daily.get(day, 0.0) + value
    wb.close()

    if not daily:
        raise RuntimeError("No usable Redsea daily sales were found")

    start, end = min(daily), max(daily)
    dates: list[date] = []
    values: list[float] = []
    cursor = start
    while cursor <= end:
        dates.append(cursor)
        values.append(max(0.0, daily.get(cursor, 0.0)))
        cursor += timedelta(days=1)
    return dates, values, source_rows


def _seasonal_naive(train: list[float], horizon: int) -> list[float]:
    mutable = list(train)
    out: list[float] = []
    for _ in range(horizon):
        pred = max(0.0, mutable[-7])
        mutable.append(pred)
        out.append(pred)
    return out


def _metrics(actual: list[float], predicted: list[float]) -> dict:
    if not actual:
        return {"wape": None, "mae": None, "rmse": None, "accuracy_proxy_pct": None}
    abs_error = sum(abs(a - p) for a, p in zip(actual, predicted))
    denom = sum(abs(a) for a in actual)
    wape = abs_error / max(denom, 1e-9)
    mae = abs_error / len(actual)
    rmse = math.sqrt(sum((a - p) ** 2 for a, p in zip(actual, predicted)) / len(actual))
    return {
        "wape": wape,
        "mae": mae,
        "rmse": rmse,
        "accuracy_proxy_pct": max(0.0, (1.0 - wape) * 100.0),
    }


def _total_wape(actual_totals: list[float], predicted_totals: list[float]) -> float | None:
    if not actual_totals:
        return None
    error = sum(abs(a - p) for a, p in zip(actual_totals, predicted_totals))
    denom = sum(abs(a) for a in actual_totals)
    return error / max(denom, 1e-9)


def evaluate(dates: list[date], values: list[float], horizon: int, step: int) -> dict:
    min_history = 56
    all_actual: list[float] = []
    all_pred: list[float] = []
    all_naive: list[float] = []
    actual_totals: list[float] = []
    pred_totals: list[float] = []
    naive_totals: list[float] = []
    coverage_hits = 0
    coverage_total = 0
    winners: Counter[str] = Counter()
    folds: list[dict] = []

    for origin in range(min_history, len(values) - horizon + 1, step):
        train = values[:origin]
        actual = values[origin: origin + horizon]
        generated = forecast(train, dates[origin - 1], horizon)
        predicted = [float(item["predicted"]) for item in generated]
        naive = _seasonal_naive(train, horizon)
        winner = str(generated[0]["model_name"])
        winners[winner] += 1

        fold_metrics = _metrics(actual, predicted)
        naive_metrics = _metrics(actual, naive)
        actual_total = sum(actual)
        pred_total = sum(predicted)
        naive_total = sum(naive)
        actual_totals.append(actual_total)
        pred_totals.append(pred_total)
        naive_totals.append(naive_total)

        for a, item in zip(actual, generated):
            coverage_total += 1
            if float(item["lower"]) <= a <= float(item["upper"]):
                coverage_hits += 1

        folds.append({
            "origin": dates[origin].isoformat(),
            "history_days": origin,
            "winner": winner,
            "adaptive_daily_wape": fold_metrics["wape"],
            "naive_daily_wape": naive_metrics["wape"],
            "actual_total": actual_total,
            "adaptive_total": pred_total,
            "naive_total": naive_total,
            "adaptive_total_abs_error_pct": abs(actual_total - pred_total) / max(abs(actual_total), 1e-9) * 100.0,
            "backtest_points": int(generated[0]["metrics"]["backtest_points"]),
        })
        all_actual.extend(actual)
        all_pred.extend(predicted)
        all_naive.extend(naive)

    adaptive = _metrics(all_actual, all_pred)
    naive = _metrics(all_actual, all_naive)
    adaptive_total_wape = _total_wape(actual_totals, pred_totals)
    naive_total_wape = _total_wape(actual_totals, naive_totals)
    return {
        "horizon_days": horizon,
        "step_days": step,
        "fold_count": len(folds),
        "evaluated_daily_points": len(all_actual),
        "adaptive": adaptive,
        "seasonal_naive_7_reference": naive,
        "horizon_total": {
            "adaptive_wape": adaptive_total_wape,
            "adaptive_accuracy_proxy_pct": max(0.0, (1.0 - adaptive_total_wape) * 100.0) if adaptive_total_wape is not None else None,
            "seasonal_naive_wape": naive_total_wape,
            "wape_delta_vs_naive_pct_points": (
                (naive_total_wape - adaptive_total_wape) * 100.0
                if adaptive_total_wape is not None and naive_total_wape is not None else None
            ),
        },
        "daily_wape_delta_vs_naive_pct_points": (
            (naive["wape"] - adaptive["wape"]) * 100.0
            if naive["wape"] is not None and adaptive["wape"] is not None else None
        ),
        "interval_coverage_pct": (coverage_hits / coverage_total * 100.0) if coverage_total else None,
        "winner_counts": dict(winners),
        "folds": folds,
    }


def main() -> None:
    source = Path(os.environ.get("REDSEA_XLSX", "/tmp/RedSea_Data_Cleaned.xlsx"))
    dates, values, source_rows = load_daily_sales(source)
    seven = evaluate(dates, values, horizon=7, step=7)
    thirty = evaluate(dates, values, horizon=30, step=14)

    payload = {
        "evaluation": "External Saudi Redsea adaptive point-forecast diagnostic",
        "scientific_status": "POST_OPEN_EXTERNAL_DIAGNOSTIC_NOT_FRESH_BLIND",
        "model_version": MODEL_VERSION,
        "source_sha256": SOURCE_SHA256,
        "source_rows": source_rows,
        "calendar_days": len(values),
        "date_start": dates[0].isoformat(),
        "date_end": dates[-1].isoformat(),
        "important_boundary": (
            "Redsea was already observed during project iteration and spans only about four months. "
            "These metrics are transfer diagnostics, not fresh blind production-validation claims."
        ),
        "horizon_7": seven,
        "horizon_30": thirty,
    }

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    (REPORT_DIR / "metrics.json").write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")

    def pct(x):
        return "n/a" if x is None else f"{x * 100:.2f}%"

    summary = f"""# Adaptive Merchant Forecast — Redsea external diagnostic

- Model: `{MODEL_VERSION}`
- Source SHA-256: `{SOURCE_SHA256}`
- Source rows read: **{source_rows}**
- Calendarized daily window: **{len(values)} days** ({dates[0]} to {dates[-1]})
- Scientific status: **POST_OPEN_EXTERNAL_DIAGNOSTIC_NOT_FRESH_BLIND**

## 7-day walk-forward

- Folds: **{seven['fold_count']}**
- Daily WAPE: **{pct(seven['adaptive']['wape'])}**
- 7-day TOTAL WAPE: **{pct(seven['horizon_total']['adaptive_wape'])}**
- 7-day TOTAL quality proxy (1-WAPE): **{seven['horizon_total']['adaptive_accuracy_proxy_pct']:.2f}%**
- Seasonal-naive TOTAL WAPE: **{pct(seven['horizon_total']['seasonal_naive_wape'])}**
- TOTAL WAPE improvement vs seasonal naive: **{seven['horizon_total']['wape_delta_vs_naive_pct_points']:.2f} percentage points**
- Prediction-interval empirical coverage: **{seven['interval_coverage_pct']:.2f}%**
- Selected-model counts: `{json.dumps(seven['winner_counts'], ensure_ascii=False)}`

## 30-day walk-forward

- Folds: **{thirty['fold_count']}**
- Daily WAPE: **{pct(thirty['adaptive']['wape'])}**
- 30-day TOTAL WAPE: **{pct(thirty['horizon_total']['adaptive_wape'])}**
- 30-day TOTAL quality proxy (1-WAPE): **{thirty['horizon_total']['adaptive_accuracy_proxy_pct']:.2f}%**
- Seasonal-naive TOTAL WAPE: **{pct(thirty['horizon_total']['seasonal_naive_wape'])}**
- TOTAL WAPE improvement vs seasonal naive: **{thirty['horizon_total']['wape_delta_vs_naive_pct_points']:.2f} percentage points**
- Prediction-interval empirical coverage: **{thirty['interval_coverage_pct']:.2f}%**
- Selected-model counts: `{json.dumps(thirty['winner_counts'], ensure_ascii=False)}`

## Boundary

Daily-value error and horizon-total error answer different questions. Sales Sentinel's decline decision is driven mainly by the total sales level over the next 7/30 days, so horizon-total WAPE is the more decision-relevant point-forecast diagnostic. This is still **not fresh blind validation** because Redsea was already inspected during development and covers only about four months.
"""
    (REPORT_DIR / "summary.md").write_text(summary, encoding="utf-8")
    print(summary)


if __name__ == "__main__":
    main()
