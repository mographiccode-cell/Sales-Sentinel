from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path


def _cell(value):
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return "" if value is None else value


def normalize_tabular_upload(path: Path) -> tuple[Path, bool]:
    """Return a CSV path suitable for the existing importer.

    CSV files pass through unchanged. XLSX files are read in streaming mode and
    converted to a sibling temporary CSV so validation and ingestion continue
    through exactly the same code path. The caller remains responsible for
    hashing/auditing the original uploaded file and deleting the generated CSV.
    """
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return path, False
    if suffix != ".xlsx":
        raise ValueError("Only CSV and XLSX files are supported")

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        first = next(rows, None)
        if first is None:
            raise ValueError("XLSX workbook is empty")

        headers = [str(value).strip() if value is not None else "" for value in first]
        non_empty_indexes = [index for index, value in enumerate(headers) if value]
        if not non_empty_indexes:
            raise ValueError("XLSX header row is empty")
        last_index = max(non_empty_indexes)
        headers = headers[: last_index + 1]
        if len([value for value in headers if value]) != len(set(value for value in headers if value)):
            raise ValueError("XLSX contains duplicate column names")

        target = path.with_name(f"{path.stem}.runtime.csv")
        with target.open("w", encoding="utf-8", newline="") as stream:
            writer = csv.writer(stream)
            writer.writerow(headers)
            for row in rows:
                values = list(row[: len(headers)])
                if not any(value not in (None, "") for value in values):
                    continue
                writer.writerow([_cell(value) for value in values])
        return target, True
    finally:
        workbook.close()
